from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, BarChart, Reference


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

DATA_PATH = PROJECT_ROOT / "data" / "processed"
OUTPUT_PATH = PROJECT_ROOT / "excel"

OUTPUT_PATH.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT_PATH / "EV_Market_Analytics.xlsx"


# ============================================================
# LOAD DATA
# ============================================================

yearly = pd.read_csv(
    DATA_PATH / "yearly_sales_analysis.csv"
)

manufacturer = pd.read_csv(
    DATA_PATH / "manufacturer_analysis.csv"
)

category = pd.read_csv(
    DATA_PATH / "category_analysis.csv"
)

charging = pd.read_csv(
    DATA_PATH / "sql_results" / "charging_analysis.csv"
)

state_manufacturers = pd.read_csv(
    DATA_PATH / "sql_results" / "manufacturers_by_state.csv"
)


# ============================================================
# CREATE EXCEL FILE
# ============================================================

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    yearly.to_excel(
        writer,
        sheet_name="Yearly Sales",
        index=False
    )

    manufacturer.to_excel(
        writer,
        sheet_name="Manufacturers",
        index=False
    )

    category.to_excel(
        writer,
        sheet_name="Categories",
        index=False
    )

    charging.to_excel(
        writer,
        sheet_name="Charging",
        index=False
    )

    state_manufacturers.to_excel(
        writer,
        sheet_name="State Analysis",
        index=False
    )


# ============================================================
# LOAD WORKBOOK
# ============================================================

wb = load_workbook(OUTPUT_FILE)


# ============================================================
# STYLE DATA SHEETS
# ============================================================

for ws in wb.worksheets:

    ws.freeze_panes = "A2"

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for column in ws.columns:

        max_length = 0

        letter = column[0].column_letter

        for cell in column:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            letter
        ].width = min(
            max_length + 2,
            35
        )


# ============================================================
# REMOVE OLD DASHBOARD
# ============================================================

if "Dashboard" in wb.sheetnames:

    del wb["Dashboard"]


# ============================================================
# CREATE DASHBOARD
# ============================================================

dashboard = wb.create_sheet(
    "Dashboard",
    0
)

dashboard["A1"] = (
    "EV MARKET & BUSINESS ANALYTICS"
)

dashboard["A1"].font = Font(
    bold=True,
    size=20
)


# ============================================================
# KPI CALCULATIONS
# ============================================================

total_sales = (
    manufacturer[
        "Total_Sales_2015_2024"
    ]
    .sum()
)

top_manufacturer = (
    manufacturer
    .sort_values(
        "Total_Sales_2015_2024",
        ascending=False
    )
    .iloc[0]["Manufacturer"]
)

top_category = (
    category
    .sort_values(
        "Total_Sales",
        ascending=False
    )
    .iloc[0]["Category"]
)

total_charging = (
    charging[
        "operational_charging_points"
    ]
    .sum()
)

top_charging_state = (
    charging
    .sort_values(
        "operational_charging_points",
        ascending=False
    )
    .iloc[0]["State"]
)


# ============================================================
# KPI CARDS
# ============================================================

dashboard["A3"] = "Total EV Sales"
dashboard["B3"] = total_sales

dashboard["D3"] = "Top Manufacturer"
dashboard["E3"] = top_manufacturer

dashboard["G3"] = "Top Category"
dashboard["H3"] = top_category

dashboard["A5"] = "Charging Points"
dashboard["B5"] = total_charging

dashboard["D5"] = "Leading Charging State"
dashboard["E5"] = top_charging_state


for cell in [
    "A3",
    "D3",
    "G3",
    "A5",
    "D5"
]:

    dashboard[cell].font = Font(
        bold=True
    )


# ============================================================
# YEARLY SALES LINE CHART
# ============================================================

line_chart = LineChart()

line_chart.title = (
    "EV Sales Trend — 2015 to 2024"
)

line_chart.y_axis.title = "EV Sales"

line_chart.x_axis.title = "Year"

line_chart.height = 8

line_chart.width = 15


# IMPORTANT:
# Column B = Total_EV_Sales
# Rows 2–11 = 2015–2024

data = Reference(
    wb["Yearly Sales"],
    min_col=2,
    min_row=1,
    max_row=11
)

categories = Reference(
    wb["Yearly Sales"],
    min_col=1,
    min_row=2,
    max_row=11
)

line_chart.add_data(
    data,
    titles_from_data=True
)

line_chart.set_categories(
    categories
)

line_chart.legend = None

dashboard.add_chart(
    line_chart,
    "A8"
)


# ============================================================
# TOP MANUFACTURER BAR CHART
# ============================================================

bar_chart = BarChart()

bar_chart.type = "bar"

bar_chart.title = (
    "Top 10 EV Manufacturers"
)

bar_chart.x_axis.title = "Total EV Sales"

bar_chart.y_axis.title = "Manufacturer"

bar_chart.height = 8

bar_chart.width = 15


# Manufacturer sheet:
# Column A = Manufacturer
# Column B = Total Sales

data = Reference(
    wb["Manufacturers"],
    min_col=2,
    min_row=1,
    max_row=11
)

categories = Reference(
    wb["Manufacturers"],
    min_col=1,
    min_row=2,
    max_row=11
)

bar_chart.add_data(
    data,
    titles_from_data=True
)

bar_chart.set_categories(
    categories
)

bar_chart.legend = None

dashboard.add_chart(
    bar_chart,
    "J8"
)


# ============================================================
# DASHBOARD COLUMN WIDTHS
# ============================================================

dashboard.column_dimensions["A"].width = 25
dashboard.column_dimensions["B"].width = 18
dashboard.column_dimensions["D"].width = 25
dashboard.column_dimensions["E"].width = 30
dashboard.column_dimensions["G"].width = 20
dashboard.column_dimensions["H"].width = 20


# ============================================================
# SAVE
# ============================================================

wb.save(OUTPUT_FILE)

print("=" * 60)
print("EXCEL DASHBOARD UPDATED")
print("=" * 60)

print(
    "File:",
    OUTPUT_FILE
)